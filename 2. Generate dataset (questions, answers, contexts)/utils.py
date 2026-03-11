"""Utilities: folder structure, .txt read/write, dataset.xlsx create/append, PDF text extraction."""

import json
import re
from pathlib import Path

import openpyxl
from openpyxl import Workbook
from openpyxl.worksheet.worksheet import Worksheet

DATASET_HEADERS = ["domain", "collection", "document", "question", "answer", "context"]


def safe_filename(name: str) -> str:
    """Replace path-unfriendly characters in document name for use in file paths."""
    return name.replace("/", "_").replace("\\", "_").strip() or "document"


def ensure_dirs(path: str | Path) -> None:
    """Create directory and parents if they do not exist."""
    Path(path).mkdir(parents=True, exist_ok=True)


def ensure_domain_collection_dirs(output_dir: str | Path, domain: str, collection: str) -> Path:
    """Ensure output_dir/domain/collection exists; return that directory path."""
    dir_path = Path(output_dir) / domain / collection
    ensure_dirs(dir_path)
    return dir_path


def document_path(output_dir: str | Path, domain: str, collection: str, document_name: str) -> Path:
    """Path to the .txt file for a document."""
    base = Path(output_dir) / domain / collection
    return base / f"{safe_filename(document_name)}.txt"


def read_document(path: str | Path) -> str:
    """Read full text from a .txt file. Returns empty string if file missing."""
    p = Path(path)
    if not p.is_file():
        return ""
    return p.read_text(encoding="utf-8")


def extract_text_from_pdf(pdf_path: str | Path) -> str:
    """Extract plain text from a PDF file. Requires pymupdf (pip install pymupdf)."""
    try:
        import fitz
    except ImportError:
        raise ImportError("PDF extraction requires pymupdf. Install with: pip install pymupdf") from None
    p = Path(pdf_path)
    if not p.is_file():
        return ""
    doc = fitz.open(p)
    try:
        parts = []
        for page in doc:
            parts.append(page.get_text())
        return "\n".join(parts)
    finally:
        doc.close()


def write_document(path: str | Path, content: str) -> None:
    """Write content to a .txt file; create parent dirs if needed."""
    p = Path(path)
    ensure_dirs(p.parent)
    p.write_text(content, encoding="utf-8")


def dataset_xlsx_path(output_dir: str | Path) -> Path:
    """Path to dataset.xlsx inside output_dir."""
    return Path(output_dir) / "dataset.xlsx"


# Control chars 0x00-0x1F except tab, newline, carriage return (invalid in Excel/XML)
_ILLEGAL_XML_CHARS = re.compile(r"[\x00-\x08\x0b\x0c\x0e-\x1f]")


def _sanitize_for_xlsx(value: str) -> str:
    """Remove characters that cause openpyxl IllegalCharacterError in worksheet cells."""
    if not isinstance(value, str):
        return value
    return _ILLEGAL_XML_CHARS.sub("", value)


def append_rows_to_dataset_xlsx(
    output_dir: str | Path,
    rows: list[tuple[str, str, str, str, str, str]],
) -> None:
    """Append rows to dataset.xlsx. Each row is (domain, collection, document, question, answer, context).
    Creates the file with headers if it does not exist.
    """
    path = dataset_xlsx_path(output_dir)
    if path.is_file():
        wb = openpyxl.load_workbook(path)
        ws = wb.active
        if ws is None:
            ws = wb.create_sheet("Sheet")
    else:
        ensure_dirs(Path(output_dir))
        wb = Workbook()
        ws = wb.active
        if ws is None:
            ws = wb.create_sheet("Sheet")
        ws.append(DATASET_HEADERS)
    assert isinstance(ws, Worksheet)
    for row in rows:
        ws.append([_sanitize_for_xlsx(cell) for cell in row])
    wb.save(path)


def classification_json_path(output_dir: str | Path) -> Path:
    """Path to dataset_classification.json for RAG metadata lookup."""
    return Path(output_dir) / "dataset_classification.json"


def append_to_classification_json(
    output_dir: str | Path,
    entries: list[dict],
) -> None:
    """Append classification entries to dataset_classification.json.
    Each entry: {question, answer, context, domain, collection, document}.
    """
    path = classification_json_path(output_dir)
    ensure_dirs(Path(output_dir))
    existing = []
    if path.is_file():
        existing = json.loads(path.read_text(encoding="utf-8"))
    existing.extend(entries)
    path.write_text(json.dumps(existing, ensure_ascii=False, indent=2), encoding="utf-8")


def metadata_mapping_json_path(output_dir: str | Path) -> Path:
    """Path to metadata_mapping.json for RAG filter keys."""
    return Path(output_dir) / "metadata_mapping.json"


def write_metadata_mapping_json(output_dir: str | Path, mapping: dict) -> None:
    """Write metadata_mapping.json: {domains: {key: display}, collections: {key: display}}."""
    path = metadata_mapping_json_path(output_dir)
    ensure_dirs(Path(output_dir))
    path.write_text(json.dumps(mapping, ensure_ascii=False, indent=2), encoding="utf-8")
